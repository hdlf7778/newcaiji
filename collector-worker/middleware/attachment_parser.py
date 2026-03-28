"""
附件解析中间件 — AttachmentParser
下载附件 → 按扩展名分发 → 提取文字/表格 → 返回结构化结果

支持格式:
- PDF  → pdfplumber
- DOC/DOCX → python-docx
- XLS/XLSX → openpyxl
"""
import io
import os
import logging
import tempfile
from dataclasses import dataclass, field

import httpx

logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB


@dataclass
class ParsedAttachment:
    file_name: str
    file_url: str
    file_type: str
    file_size: int = 0
    parsed_text: str = ""
    parsed_tables: list = field(default_factory=list)
    local_path: str = ""
    error: str = ""


class AttachmentParser:

    def __init__(self, save_dir: str = None):
        self.save_dir = save_dir or tempfile.mkdtemp(prefix="collector_att_")
        os.makedirs(self.save_dir, exist_ok=True)

    async def download_and_parse(self, file_url: str, file_name: str = "",
                                  client: httpx.AsyncClient = None) -> ParsedAttachment:
        """下载附件并解析内容"""
        file_type = self._detect_type(file_url, file_name)
        result = ParsedAttachment(
            file_name=file_name or file_url.split('/')[-1].split('?')[0],
            file_url=file_url,
            file_type=file_type,
        )

        # 下载
        try:
            if client is None:
                async with httpx.AsyncClient(verify=False, timeout=30) as c:
                    resp = await c.get(file_url)
            else:
                resp = await client.get(file_url)

            if resp.status_code != 200:
                result.error = f"下载失败: HTTP {resp.status_code}"
                return result

            content = resp.content
            result.file_size = len(content)

            if result.file_size > MAX_FILE_SIZE:
                result.error = f"文件过大: {result.file_size / 1024 / 1024:.1f}MB > {MAX_FILE_SIZE / 1024 / 1024}MB"
                return result

            # 保存本地
            local_path = os.path.join(self.save_dir, result.file_name)
            with open(local_path, 'wb') as f:
                f.write(content)
            result.local_path = local_path

        except Exception as e:
            result.error = f"下载异常: {e}"
            return result

        # 解析
        try:
            if file_type == 'pdf':
                result.parsed_text, result.parsed_tables = self._parse_pdf(content)
            elif file_type in ('doc', 'docx'):
                result.parsed_text = self._parse_docx(content)
            elif file_type in ('xls', 'xlsx'):
                result.parsed_text, result.parsed_tables = self._parse_xlsx(content)
            else:
                result.error = f"不支持的格式: {file_type}"
        except Exception as e:
            result.error = f"解析异常: {e}"
            logger.warning("附件解析失败 %s: %s", result.file_name, e)

        logger.info("附件解析 %s type=%s size=%dKB text=%d字 tables=%d",
                     result.file_name[:30], file_type, result.file_size // 1024,
                     len(result.parsed_text), len(result.parsed_tables))
        return result

    def _parse_pdf(self, content: bytes) -> tuple[str, list]:
        """PDF → 文字 + 表格"""
        import pdfplumber
        text_parts = []
        tables = []

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
                page_tables = page.extract_tables()
                for t in page_tables:
                    tables.append(t)

        return '\n'.join(text_parts), tables

    def _parse_docx(self, content: bytes) -> str:
        """DOCX → 文字"""
        from docx import Document
        doc = Document(io.BytesIO(content))
        return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())

    def _parse_xlsx(self, content: bytes) -> tuple[str, list]:
        """XLSX → 文字 + 表格"""
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        text_parts = []
        tables = []

        for ws in wb.worksheets:
            sheet_rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else '' for c in row]
                if any(cells):
                    sheet_rows.append(cells)
                    text_parts.append(' '.join(c for c in cells if c))
            if sheet_rows:
                tables.append(sheet_rows)

        wb.close()
        return '\n'.join(text_parts), tables

    @staticmethod
    def _detect_type(url: str, name: str = "") -> str:
        for s in [name.lower(), url.lower()]:
            if s.endswith('.pdf'):
                return 'pdf'
            if s.endswith('.docx'):
                return 'docx'
            if s.endswith('.doc'):
                return 'doc'
            if s.endswith('.xlsx'):
                return 'xlsx'
            if s.endswith('.xls'):
                return 'xls'
        return 'unknown'
